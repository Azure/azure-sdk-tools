"""Feedback workflow service.

Processes user feedback as a plain Python workflow (no LLM agent).
Saves feedback records to Azure Blob Storage as Excel files (monthly),
mirroring the Go backend approach.
"""

from __future__ import annotations

import json
import logging
from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook

from config.app_config import get as cfg
from models.feedback import FeedbackRequest, FeedbackResponse, Reaction
from utils.azure_storage import download_blob, upload_blob

logger = logging.getLogger(__name__)

_SHEET_NAME = "Feedback"
_HEADERS = [
    "Timestamp",
    "TenantID",
    "Messages",
    "Reaction",
    "Comment",
    "Reasons",
    "Link",
    "ChannelID",
    "UserName",
]


class FeedbackService:
    """Workflow that persists feedback and creates GitHub issues for bad cases."""

    async def process(self, req: FeedbackRequest) -> FeedbackResponse:
        """Run the full feedback workflow.

        Steps:
          1. Save the feedback record to Azure Blob Storage.
          2. If reaction is "bad", create a GitHub issue.
        """
        result = FeedbackResponse()

        await self._save_feedback(req)
        result.saved = True

        if req.reaction == Reaction.bad:
            result.issue_url = await self._create_github_issue(req)

        return result

    async def _save_feedback(self, req: FeedbackRequest) -> None:
        """Save feedback record to Azure Blob Storage.

        Appends the entry to a monthly feedback file (feedback_YYYY_MM.xlsx)
        in the configured storage container.
        """
        container = cfg("STORAGE_FEEDBACK_CONTAINER", "")
        if not container:
            raise RuntimeError("STORAGE_FEEDBACK_CONTAINER not configured")

        now = datetime.now(timezone.utc)
        filename = f"feedback_{now.year:04d}_{now.month:02d}.xlsx"

        # Try to download existing Excel file
        existing_data = await download_blob(container, filename)

        if existing_data:
            wb = load_workbook(filename=BytesIO(existing_data))
            ws = wb[_SHEET_NAME] if _SHEET_NAME in wb.sheetnames else wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = _SHEET_NAME
            ws.append(_HEADERS)

        # Build the new row matching Go backend column order
        row = [
            now.isoformat(),
            req.tenant_id,
            "",  # Messages column is not used in the current workflow
            req.reaction,
            req.comment or "",
            json.dumps(req.reasons, ensure_ascii=False),
            req.link or "",
            req.channel_id or "",
            req.user_name or "",
        ]
        ws.append(row)

        # Write workbook to bytes and upload
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        await upload_blob(container, filename, buf.getvalue())
        logger.info("Saved feedback to %s/%s", container, filename)

    async def _create_github_issue(self, req: FeedbackRequest) -> Optional[str]:
        """Create a GitHub issue for a bad feedback case.

        Returns the issue URL on success, or None on failure.
        """
        # TODO: implement with GitHub API
        return None
