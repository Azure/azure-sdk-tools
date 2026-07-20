"""Tools for the TypeSpec Authoring Skill Self-Evolving Agent.

These functions are registered with the agent (via ``agent_framework``) and
implement the concrete, side-effecting actions behind the **benchmark-gated**
workflow (the draft PR is opened last, only if the benchmark clears its
pass-rate threshold — see ``workflow.py``):

    Step 1  Analyze telemetry + update skill -> read_prompt_excel (cluster common
                                             use cases from user telemetry) +
                                             fetch_documentation (grounding); the
                                             agent updates
                                             references/reference-document-links.md,
                                             and AVOIDS touching other skill markdown
                                             files unless truly needed and necessary
                                             (then minimal), then
                                             github_tools.push_skill_changes commits
                                             the edits to a branch
    Step 2  Run the benchmark test        -> run_benchmark (local) / ADO 8178 (remote)
    Step 3  Score the benchmark           -> compute_pass_rate (the numeric PR gate)
                                             + summarize_benchmark_results (report)
    Step 4  Analyze current document gaps -> (reasoning over the above outputs)
    GATE    Open a draft PR iff it passed -> open_draft_pr_if_benchmark_passed:
                                             when the benchmark pass rate EXCEEDS the
                                             threshold (default 75%) it opens a *draft*
                                             PR for the skill update via
                                             github_tools.open_draft_pr; otherwise it
                                             opens no PR and reports why.

All heavy lifting that needs the repository (Vally, the skill sources) runs
through ``run_benchmark`` against a working directory that is mounted or cloned
into the container. Documentation grounding is fully self-contained via httpx,
and ``read_prompt_excel`` parses a workbook of real user prompts via openpyxl.
"""

from __future__ import annotations

import json
import logging
import os
import re
import subprocess
from html.parser import HTMLParser
from pathlib import Path
from typing import Annotated, Optional
from urllib.parse import urlparse

import httpx

logger = logging.getLogger(__name__)

_FETCH_TIMEOUT_SECONDS = 15
_MAX_DOC_CHARS = 12000
# Where the azure-sdk-tools checkout lives inside the container / host. The
# evaluate directory holds .vally.yaml and the eval suites.
_EVALUATE_DIR_ENV = "TYPESPEC_EVALUATE_DIR"
_DEFAULT_EVALUATE_DIR = (
    ".github/skills/azure-typespec-author/evaluate"
)


class _TextExtractor(HTMLParser):
    """Extract visible text from HTML, skipping script/style/nav noise."""

    _SKIP = frozenset({"script", "style", "noscript", "nav", "footer", "header", "svg", "head"})

    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []
        self._skip = 0

    def handle_starttag(self, tag, attrs):
        if tag in self._SKIP:
            self._skip += 1

    def handle_endtag(self, tag):
        if tag in self._SKIP and self._skip > 0:
            self._skip -= 1

    def handle_data(self, data):
        if self._skip == 0:
            s = data.strip()
            if s:
                self.parts.append(s)

    def text(self) -> str:
        return re.sub(r"\s+", " ", " ".join(self.parts)).strip()


def _is_public_http_url(url: str) -> bool:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"}:
        return False
    host = (parsed.hostname or "").lower()
    return bool(host) and host not in {"localhost", "127.0.0.1", "::1"}


async def fetch_documentation(
    url: Annotated[str, "Public https URL of the authoritative TypeSpec/Azure doc to ground an edit in."],
    max_chars: Annotated[int, "Maximum characters of extracted text to return."] = 8000,
) -> str:
    """Fetch a public documentation page and return its extracted text.

    Use this in Step 1 to ground every skill edit in content you actually
    fetched, and in Step 4 to confirm whether a page contains the guidance a
    failing benchmark case needs. Never author from memory alone.
    """
    if not _is_public_http_url(url):
        raise ValueError("Only public http/https URLs are allowed.")
    bounded = max(1000, min(int(max_chars), _MAX_DOC_CHARS))
    headers = {
        "User-Agent": "typespec-skill-self-evolving-agent/1.0 (+https://github.com/Azure/azure-sdk-tools)",
        "Accept": "text/html,application/xhtml+xml,text/markdown,*/*;q=0.8",
    }
    async with httpx.AsyncClient(headers=headers, follow_redirects=True, timeout=_FETCH_TIMEOUT_SECONDS) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        content_type = resp.headers.get("content-type", "")
        raw = resp.text
    if "html" in content_type.lower():
        extractor = _TextExtractor()
        extractor.feed(raw)
        body = extractor.text()
    else:
        body = re.sub(r"\s+", " ", raw).strip()
    return f"# Source: {url}\n\n{body[:bounded]}"


def _evaluate_dir() -> Path:
    return Path(os.environ.get(_EVALUATE_DIR_ENV, _DEFAULT_EVALUATE_DIR)).resolve()


def run_benchmark(
    suite: Annotated[str, "Vally suite to run, e.g. 'armtemplate-forced', 'versioning-trigger', or a full mode like 'forced'."],
    skill_dir: Annotated[str, "Skill dir passed to Vally. Use '..' for candidate/baseline skill, '/tmp/no-skills' for no-skill runs."] = "..",
) -> str:
    """Run one Vally benchmark suite locally and return the raw stdout/stderr.

    **LOCAL MODE ONLY.** This shells out to the Node ``vally`` CLI and needs an
    azure-sdk-tools checkout (``TYPESPEC_EVALUATE_DIR``). The Foundry hosted
    container has neither, so this returns an error there — run the benchmark
    remotely instead by opening a PR (which triggers ``skill-eval`` CI) or via
    ``dispatch_workflow`` + ``get_latest_workflow_run`` from ``github_tools``.

    Executes ``vally eval --suite <suite> --skill-dir <skill_dir>`` from the
    evaluate directory. Run this for both the candidate change and the pre-change
    baseline so a delta can be computed in Step 3.
    """
    evaluate_dir = _evaluate_dir()
    if not (evaluate_dir / ".vally.yaml").exists():
        return (
            f"ERROR: Vally config not found at {evaluate_dir}. This is local-only; set "
            f"{_EVALUATE_DIR_ENV} to the azure-typespec-author/evaluate directory of an "
            "azure-sdk-tools checkout, or run the benchmark remotely via the GitHub "
            "tools (open_skill_pull_request triggers skill-eval CI; poll with "
            "get_latest_workflow_run)."
        )
    cmd = [
        "vally", "eval",
        "--suite", suite,
        "--skill-dir", skill_dir,
        "--output-dir", "./result",
        "--workspace", "./debug",
        "--verbose",
    ]
    logger.info("Running benchmark: %s (cwd=%s)", " ".join(cmd), evaluate_dir)
    try:
        proc = subprocess.run(
            cmd, cwd=str(evaluate_dir), capture_output=True, text=True, timeout=3600
        )
    except FileNotFoundError:
        return "ERROR: 'vally' CLI not found on PATH. Install @microsoft/vally-cli."
    except subprocess.TimeoutExpired:
        return f"ERROR: benchmark suite '{suite}' timed out after 3600s."
    tail = (proc.stdout or "")[-6000:]
    err = (proc.stderr or "")[-2000:]
    return f"exit_code={proc.returncode}\n--- stdout (tail) ---\n{tail}\n--- stderr (tail) ---\n{err}"


def _count_from_markdown(text: str) -> Optional[tuple[int, int, list[str]]]:
    """Count passed/failed stimuli from a Vally ``eval-results.md`` summary table.

    The CI benchmark publishes its per-stimulus verdicts as a markdown table with a
    ``Pass Rate`` column (cells like ``0/1`` / ``1/1``) and a ``Stimulus`` column.
    Returns ``(passed, failed, failing_case_names)`` or ``None`` if no such table is
    found (so the caller can fall back to JSONL parsing).
    """
    lines = text.splitlines()
    passed = 0
    total = 0
    cases: list[str] = []
    stim_idx: Optional[int] = None
    rate_idx: Optional[int] = None
    found_table = False
    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("|"):
            # A blank/non-table line ends the current table's column mapping.
            if stripped == "":
                stim_idx = rate_idx = None
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        lowered = [c.lower() for c in cells]
        if "pass rate" in lowered:  # header row
            rate_idx = lowered.index("pass rate")
            stim_idx = lowered.index("stimulus") if "stimulus" in lowered else 0
            continue
        if set("".join(cells)) <= set("-: "):  # separator row (---)
            continue
        if rate_idx is None or rate_idx >= len(cells):
            continue
        m = re.match(r"(\d+)\s*/\s*(\d+)", cells[rate_idx])
        if not m:
            continue
        found_table = True
        p, t = int(m.group(1)), int(m.group(2))
        passed += p
        total += t
        if p < t:
            name = cells[stim_idx] if stim_idx is not None and stim_idx < len(cells) else "?"
            cases.append(name[:120])
    if not found_table:
        return None
    return passed, total - passed, cases


def _count_pass_fail(text: str) -> tuple[int, int, list[str]]:
    """Count passed/failed cases in Vally results text.

    Prefers the ``eval-results.md`` markdown summary table (what the CI benchmark
    publishes); falls back to parsing ``results.jsonl`` rows (local mode).
    Returns ``(passed, failed, failing_case_names)``. Shared by
    ``summarize_benchmark_results`` and ``compute_pass_rate``.
    """
    md = _count_from_markdown(text)
    if md is not None and (md[0] + md[1]) > 0:
        return md
    passed = 0
    failed = 0
    cases: list[str] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if not isinstance(row, dict):
            continue
        # Skip the trailing run-summary line Vally emits.
        if row.get("type") == "run-summary":
            continue
        grade = row.get("gradeResult") if isinstance(row.get("gradeResult"), dict) else {}
        name = (
            grade.get("stimulusName")
            or row.get("stimulus")
            or row.get("name")
            or row.get("itemId")
            or "?"
        )
        ok = grade.get("passed") if "passed" in grade else row.get("passed")
        score = row.get("score")
        if ok is None and score is not None:
            ok = bool(score)
        if ok:
            passed += 1
        else:
            failed += 1
            cases.append(str(name))
    return passed, failed, cases


def _read_results_text(results_path: str, content: str) -> str | None:
    """Resolve results text from raw ``content`` (preferred) or a file path."""
    if content and content.strip():
        return content
    if not results_path:
        return None
    path = Path(results_path)
    if not path.exists():
        return None
    return path.read_text(encoding="utf-8")


def compute_pass_rate(
    results_path: Annotated[str, "Path to a Vally results.jsonl file (local mode)."] = "",
    content: Annotated[str, "Raw results.jsonl text (remote/CI mode). Takes precedence over results_path."] = "",
) -> str:
    """Compute the numeric benchmark pass rate as a JSON object.

    Returns ``{"status":"ok","cases":N,"passed":P,"failed":F,"pass_rate":R}`` where
    ``pass_rate`` is a float percentage (0-100). Used to **gate** the draft PR:
    the workflow opens the PR only when ``pass_rate`` clears its threshold (e.g.
    75%). Accepts the same inputs as ``summarize_benchmark_results``.
    """
    text = _read_results_text(results_path, content)
    if text is None:
        return json.dumps(
            {"status": "error", "message": f"provide 'content' or a valid 'results_path' (got: {results_path!r})."},
            ensure_ascii=False,
        )
    passed, failed, _ = _count_pass_fail(text)
    total = passed + failed
    rate = round(passed / total * 100, 1) if total else 0.0
    return json.dumps(
        {"status": "ok", "cases": total, "passed": passed, "failed": failed, "pass_rate": rate},
        ensure_ascii=False,
    )


# Default benchmark pass-rate percentage a candidate must EXCEED to open a PR.
_DEFAULT_MIN_PASS_RATE = 75.0


def open_draft_pr_if_benchmark_passed(
    branch: Annotated[str, "Head branch already pushed via push_skill_changes (carries the skill edits)."],
    title: Annotated[str, "Draft PR title."],
    body: Annotated[str, "Draft PR body — include the benchmark test report."],
    results_path: Annotated[str, "Path to a Vally results.jsonl file (local mode)."] = "",
    content: Annotated[str, "Raw results.jsonl text (remote/CI mode). Takes precedence over results_path."] = "",
    min_pass_rate: Annotated[float, "Pass-rate percentage the benchmark must EXCEED to open the PR (default 75)."] = _DEFAULT_MIN_PASS_RATE,
    base: Annotated[str, "Base branch the PR targets."] = "main",
) -> str:
    """**Benchmark gate:** open a DRAFT PR for the skill update **only if** the benchmark passed.

    This is the final, deterministic step of the self-evolving workflow. It scores
    the Vally results (same inputs as ``compute_pass_rate``) and decides:

    * ``pass_rate > min_pass_rate`` (e.g. **> 75%**) → opens a **draft** PR from
      ``branch`` via ``github_tools.open_draft_pr`` and returns the PR info
      (``pr_number``, ``html_url``, ``draft``).
    * ``pass_rate <= min_pass_rate`` → opens **no** PR and returns a ``skipped``
      result with the reason and the failing cases, so a human can inspect the
      branch instead.
    * no gradable results → returns an ``error`` (the gate cannot be evaluated).

    The PR is always a draft and is never merged — a human promotes it out of draft
    after reviewing the benchmark evidence embedded in ``body``. Returns a JSON
    object describing the decision.
    """
    text = _read_results_text(results_path, content)
    if text is None:
        return json.dumps(
            {"status": "error",
             "action": "gate_not_evaluated",
             "message": f"provide 'content' or a valid 'results_path' (got: {results_path!r}); "
                        "cannot evaluate the benchmark gate."},
            ensure_ascii=False,
        )

    passed, failed, failing = _count_pass_fail(text)
    total = passed + failed
    if total == 0:
        return json.dumps(
            {"status": "error",
             "action": "gate_not_evaluated",
             "message": "results contained no gradable cases; the benchmark gate was not evaluated."},
            ensure_ascii=False,
        )
    rate = round(passed / total * 100, 1)

    # Gate not cleared: do NOT open a PR (strictly-greater-than threshold).
    if rate <= min_pass_rate:
        return json.dumps(
            {
                "status": "skipped",
                "action": "no_pr",
                "reason": f"pass_rate {rate}% did not exceed the {min_pass_rate}% threshold",
                "pass_rate": rate,
                "min_pass_rate": min_pass_rate,
                "cases": total,
                "passed": passed,
                "failed": failed,
                "failing_cases": failing[:50],
                "branch": branch,
            },
            ensure_ascii=False,
        )

    # Gate cleared → open the draft PR. Lazy import keeps the scoring helpers usable
    # without pulling in the GitHub HTTP client, and avoids any import cycle.
    from . import github_tools

    pr_raw = github_tools.open_draft_pr(branch=branch, title=title, body=body, base=base)
    try:
        pr = json.loads(pr_raw)
    except json.JSONDecodeError:
        pr = {"status": "error", "message": pr_raw}

    if pr.get("status") != "ok":
        return json.dumps(
            {
                "status": "error",
                "action": "pr_failed",
                "message": f"benchmark passed ({rate}%) but opening the draft PR failed: "
                           f"{pr.get('message')}",
                "pass_rate": rate,
                "min_pass_rate": min_pass_rate,
                "branch": branch,
            },
            ensure_ascii=False,
        )

    return json.dumps(
        {
            "status": "ok",
            "action": "draft_pr_opened",
            "pass_rate": rate,
            "min_pass_rate": min_pass_rate,
            "cases": total,
            "passed": passed,
            "failed": failed,
            "pr_number": pr.get("pr_number"),
            "html_url": pr.get("html_url"),
            "draft": pr.get("draft", True),
            "branch": branch,
        },
        ensure_ascii=False,
    )


def _sharepoint_download_url(url: str) -> str:
    """Best-effort: turn a SharePoint/OneDrive *sharing* link into a direct
    download URL by appending ``download=1``. Auth is still required for private
    tenants, so this only helps for anonymously shared or already-authorized links.
    """
    lower = url.lower()
    if "sharepoint.com" in lower or "1drv.ms" in lower or "-my.sharepoint" in lower:
        sep = "&" if "?" in url else "?"
        if "download=1" not in lower:
            return f"{url}{sep}download=1"
    return url


def _run_async(coro):
    """Run an async coroutine from sync tool code.

    Uses ``asyncio.run`` when no loop is running; if a loop is already running
    in this thread, executes the coroutine in a dedicated thread with its own
    loop so we never call ``asyncio.run`` re-entrantly.
    """
    import asyncio

    try:
        asyncio.get_running_loop()
    except RuntimeError:
        return asyncio.run(coro)

    import concurrent.futures

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
        return ex.submit(lambda: asyncio.run(coro)).result()


def _is_sharepoint_url(url: str) -> bool:
    """True for SharePoint / OneDrive (personal or org) sharing links."""
    host = (urlparse(url).hostname or "").lower()
    return (
        host.endswith("sharepoint.com")
        or host.endswith("-my.sharepoint.com")
        or host == "1drv.ms"
        or host.endswith(".1drv.ms")
    )


def _graph_share_id(url: str) -> str:
    """Encode a sharing URL into a Graph ``shares`` id (``u!<base64url>``).

    See https://learn.microsoft.com/graph/api/shares-get — a share id is
    ``u!`` + unpadded base64url of the UTF-8 URL.
    """
    import base64

    b64 = base64.urlsafe_b64encode(url.encode("utf-8")).decode("utf-8").rstrip("=")
    return f"u!{b64}"


def _graph_token() -> str:
    """Acquire a Microsoft Graph bearer token.

    Resolution order:

    1. An explicit ``GRAPH_TOKEN`` (env or ``.env``) — the simplest local escape
       hatch: paste a token from Graph Explorer or ``Connect-MgGraph`` (needs the
       ``Files.Read.All`` / ``Sites.Read.All`` delegated scope).
    2. ``DefaultAzureCredential`` with the ``https://graph.microsoft.com/.default``
       scope — the hosted path: the container's managed identity must be granted
       the ``Files.Read.All`` **application** permission on Microsoft Graph.

    Note: a plain ``az login`` identity usually does **not** work — the Azure CLI
    first-party app's Graph token does not carry ``Files.Read``/``Sites.Read``.
    """
    explicit = (os.environ.get("GRAPH_TOKEN") or "").strip()
    if explicit:
        return explicit.removeprefix("Bearer ").strip()

    from azure.identity import DefaultAzureCredential

    return DefaultAzureCredential().get_token("https://graph.microsoft.com/.default").token


def _download_sharepoint_via_graph(url: str) -> bytes:
    """Download a SharePoint/OneDrive shared file's bytes via Microsoft Graph.

    Resolves the sharing link through the ``/shares/{id}/driveItem/content``
    endpoint (which returns the file content, following the download redirect)
    using a Graph token from :func:`_graph_token`. The identity behind the token
    must have access to the shared item and the ``Files.Read.All`` /
    ``Sites.Read.All`` scope.

    Raises on any auth/transport error so the caller can surface a clear message.
    """
    token = _graph_token()
    share_id = _graph_share_id(url)
    graph_url = (
        f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem/content"
    )
    headers = {"Authorization": f"Bearer {token}"}
    with httpx.Client(follow_redirects=True, timeout=_FETCH_TIMEOUT_SECONDS) as client:
        resp = client.get(graph_url, headers=headers)
        resp.raise_for_status()
        return resp.content


def read_prompt_excel(
    source: Annotated[
        str,
        "Local path to an .xlsx file, an https URL that returns the .xlsx bytes, or a "
        "SharePoint/OneDrive sharing link. SharePoint/OneDrive links are resolved via "
        "Microsoft Graph using the ambient identity (az login / managed identity), which "
        "must have access to the shared file.",
    ],
    column: Annotated[
        str,
        "Optional header name of the column holding the user prompt/question (case-insensitive, "
        "substring match). If empty, all non-empty cell text is returned.",
    ] = "",
    max_rows: Annotated[int, "Maximum number of prompt rows to return."] = 500,
) -> str:
    """Read an Excel workbook of **user telemetry** (prompts) and return the rows.

    Step 1 grounding for the self-evolving workflow: the agent consumes a workbook
    of real user prompts/questions (user telemetry, e.g. exported via WorkIQ), then
    reasons over the returned rows to find **common use cases** and map them onto
    the skill's case categories. A use case counts as **common only when similar
    prompts appear at least 5 times** in the telemetry; the agent should **only
    update ``references/reference-document-links.md`` for a common case that is not
    already covered**, ignoring rare/one-off prompts (fewer than 5 similar
    occurrences). It **avoids**
    touching other skill markdown files unless truly needed and necessary (then only
    minimal edits). Returns a JSON object with the detected ``column``,
    ``row_count``, and a ``rows`` list of prompt strings (deduplicated, truncated
    to ``max_rows``).

    Reads ``.xlsx`` via ``openpyxl``. ``source`` may be a local path, a direct
    download URL, or a **SharePoint/OneDrive sharing link** (resolved through the
    Microsoft Graph ``shares`` API using the ambient identity, which must have
    access to the shared file).
    """
    try:
        from openpyxl import load_workbook
    except Exception:  # pragma: no cover
        return json.dumps(
            {"status": "error", "message": "openpyxl is not installed; add openpyxl to requirements."},
            ensure_ascii=False,
        )

    import io

    try:
        if _is_sharepoint_url(source):
            # SharePoint/OneDrive sharing links are auth-gated — a plain GET
            # returns 401. Read them online via the Foundry toolbox (WorkIQ),
            # which reads the document on the caller's behalf; fall back to
            # Microsoft Graph (needs Files.Read.All) if the toolbox is unavailable.
            data = None
            toolbox_err = None
            try:
                from tools.foundry_toolbox_tools import fetch_sharepoint_file_bytes

                data = io.BytesIO(_run_async(fetch_sharepoint_file_bytes(source)))
            except Exception as exc:
                toolbox_err = f"{type(exc).__name__}: {str(exc)[:200]}"
                logger.info("Toolbox SharePoint read failed (%s); trying Graph.", toolbox_err)
            if data is None:
                try:
                    data = io.BytesIO(_download_sharepoint_via_graph(source))
                except Exception as exc:
                    return json.dumps(
                        {
                            "status": "error",
                            "message": (
                                "Could not read the SharePoint/OneDrive workbook. Foundry "
                                f"toolbox: {toolbox_err or 'not configured'}. Microsoft Graph: "
                                f"{type(exc).__name__}: {str(exc)[:160]}. Fixes: configure the "
                                "Foundry toolbox (FOUNDRY_TOOLBOX_MCP_URL) and grant its identity "
                                "access; or set GRAPH_TOKEN with Files.Read.All; or download the "
                                "workbook and pass a local file path."
                            ),
                        },
                        ensure_ascii=False,
                    )
        elif _is_public_http_url(source):
            dl_url = _sharepoint_download_url(source)
            with httpx.Client(follow_redirects=True, timeout=_FETCH_TIMEOUT_SECONDS) as client:
                resp = client.get(dl_url)
                resp.raise_for_status()
                ctype = resp.headers.get("content-type", "").lower()
                if "html" in ctype:
                    return json.dumps(
                        {"status": "error",
                         "message": "URL returned HTML, not an .xlsx (likely a SharePoint auth/redirect "
                                    "page). Download the workbook and pass a local file path instead."},
                        ensure_ascii=False,
                    )
                data = io.BytesIO(resp.content)
        else:
            p = Path(source)
            if not p.exists():
                return json.dumps({"status": "error", "message": f"File not found: {source}"}, ensure_ascii=False)
            data = io.BytesIO(p.read_bytes())

        wb = load_workbook(data, read_only=True, data_only=True)
    except Exception as exc:
        return json.dumps(
            {"status": "error", "message": f"Failed to open workbook: {type(exc).__name__}: {str(exc)[:200]}"},
            ensure_ascii=False,
        )

    rows: list[str] = []
    seen: set[str] = set()
    detected_col = column or ""
    try:
        for ws in wb.worksheets:
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            col_idx = None
            if column and header:
                target = column.strip().lower()
                for i, h in enumerate(header):
                    if h and target in str(h).strip().lower():
                        col_idx = i
                        detected_col = str(h)
                        break
            # If a header row exists but no column matched and none requested,
            # still scan all cells (col_idx stays None => take all text).
            for r in it:
                if r is None:
                    continue
                if col_idx is not None:
                    vals = [r[col_idx]] if col_idx < len(r) else []
                else:
                    vals = list(r)
                for v in vals:
                    if v is None:
                        continue
                    s = str(v).strip()
                    if not s or s.lower() in seen:
                        continue
                    seen.add(s.lower())
                    rows.append(s)
                    if len(rows) >= max_rows:
                        raise StopIteration
    except StopIteration:
        pass
    finally:
        wb.close()

    return json.dumps(
        {"status": "ok", "column": detected_col, "row_count": len(rows), "rows": rows},
        ensure_ascii=False,
    )


def summarize_benchmark_results(
    results_path: Annotated[
        str,
        "Path to a Vally results.jsonl file (local mode). Leave empty and pass 'content' when summarizing a downloaded CI artifact.",
    ] = "",
    content: Annotated[
        str,
        "Raw results.jsonl text (remote mode) — e.g. the 'content' returned by download_workflow_artifact. Takes precedence over results_path.",
    ] = "",
) -> str:
    """Summarize Vally results.jsonl into per-case pass/fail counts.

    Works in both modes: pass ``results_path`` for a local file, or ``content``
    for the raw JSONL text downloaded from a CI artifact
    (``download_workflow_artifact``). Use in Step 3 to build the report; compare
    a candidate summary against a baseline summary to compute the delta and flag
    regressions.
    """
    text = _read_results_text(results_path, content)
    if text is None:
        return f"ERROR: provide 'content' or a valid 'results_path' (got: {results_path!r})."
    passed, failed, cases = _count_pass_fail(text)
    total = passed + failed
    rate = f"{(passed / total * 100):.1f}%" if total else "n/a"
    lines = [f"cases={total} passed={passed} failed={failed} pass_rate={rate}"]
    if cases:
        lines.append("failing_cases:")
        lines.extend(f"  - {c}" for c in cases[:50])
    return "\n".join(lines)
